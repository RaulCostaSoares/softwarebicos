#!/usr/bin/env python3
"""
Extrai presets de aeronave e referencia de fluxometro do arquivo
"CALIBRACAO AERONAVE E FLUXOMETRO.xlsx".

Uso:
  python scripts/importar_aeronave_fluxometro_xlsx.py --input "C:\\caminho\\arquivo.xlsx" --output js/data/aeronaves_fluxometro.json
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def as_num(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip().replace(",", ".")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def extract_models(ws, model_col: int, vmax_col: int, vmed_col: int, vmin_col: int) -> list[dict[str, Any]]:
    found: list[dict[str, Any]] = []
    for row in range(1, ws.max_row + 1):
        model = as_text(ws.cell(row=row, column=model_col).value)
        if not model or model.upper() in {"MODEL", "CALCULADORA"}:
            continue
        vmax = as_num(ws.cell(row=row, column=vmax_col).value)
        vmed = as_num(ws.cell(row=row, column=vmed_col).value)
        vmin = as_num(ws.cell(row=row, column=vmin_col).value)
        if vmax is None or vmed is None or vmin is None:
            continue
        found.append(
            {
                "modelo": model,
                "vmaxKmh": vmax,
                "vmedKmh": vmed,
                "vminKmh": vmin,
            }
        )
    return found


def dedupe_models(*model_lists: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_name: dict[str, dict[str, Any]] = {}
    for models in model_lists:
        for m in models:
            key = m["modelo"].upper()
            if key not in by_name:
                by_name[key] = m
    return sorted(by_name.values(), key=lambda x: x["modelo"])


def extract_fluxometer_params(ws) -> dict[str, Any]:
    faixa_min = as_num(ws.cell(row=11, column=2).value)
    taxa_min = as_num(ws.cell(row=12, column=2).value)
    faixa_max = as_num(ws.cell(row=15, column=2).value)
    taxa_max = as_num(ws.cell(row=16, column=2).value)
    return {
        "faixaMinM": faixa_min,
        "taxaMinLHa": taxa_min,
        "faixaMaxM": faixa_max,
        "taxaMaxLHa": taxa_max,
        "formula": "vazao_total_l_min = (velocidade_kmh * faixa_m * taxa_l_ha) / 600",
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="Arquivo XLSX origem")
    parser.add_argument("--output", required=True, help="Arquivo JSON de saida")
    args = parser.parse_args()

    in_path = Path(args.input)
    if not in_path.exists():
        raise SystemExit(f"Arquivo nao encontrado: {in_path}")

    wb = load_workbook(str(in_path), data_only=True, read_only=True)
    if len(wb.worksheets) < 2:
        wb.close()
        raise SystemExit("Arquivo nao possui as abas esperadas.")

    ws_aero = wb.worksheets[0]
    ws_flux = wb.worksheets[1]

    models_aero = extract_models(ws_aero, model_col=13, vmax_col=14, vmed_col=15, vmin_col=16)
    models_flux = extract_models(ws_flux, model_col=11, vmax_col=12, vmed_col=13, vmin_col=14)
    models = dedupe_models(models_aero, models_flux)

    payload = {
        "fonteArquivo": in_path.name,
        "geradoEmUtc": datetime.now(timezone.utc).isoformat(),
        "aeronaves": models,
        "fluxometroReferencia": extract_fluxometer_params(ws_flux),
    }

    wb.close()

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Arquivo gerado: {out_path} ({len(models)} aeronaves)")


if __name__ == "__main__":
    main()

